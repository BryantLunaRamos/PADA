"""
Bryant Luna-Ramos
6/16/26
"""

import argparse
import csv
import os
import re
import sqlite3
from difflib import SequenceMatcher

import common
import matplotlib.pyplot as plt
import openpyxl

DIIT_KEYWORDS = [
    "DIIT", "information technology", "instructional technology", "technology",
    "software", "hardware", "network", "server", "laptop", "desktop", "tablet",
    "chromebook", "wireless", "data center", "cabling", "IT services",
    "IT support", "IT consulting", "cloud", "digital", "computer", "device",
    "cyber", "telecommunications", "infrastructure", "system",
]

DIIT_EXCLUDE_PHRASES = [
    "family child care", "crisis management system", "system-wide", "systemwide",
    "system wide", "fire alarm", "fire suppression", "sprinkler", "standpipe",
    "security system", "hvac", "air condition", "boiler", "plumbing", "backflow",
    "fuel oil", "public address system", "gas leak detection", "de-watering",
    "kitchen exhaust", "water treatment", "direct digital control",
    "window shades", "legal process server",
    "vendor does not have order in system", "doc posted in city",
]

TABLE_CONFIGS = {
    "contracts_registered": ["is_diit INTEGER DEFAULT 0"],
    "contracts_pending":    ["is_diit INTEGER DEFAULT 0"],
}

GENERIC_SUFFIXES = {
    "INC", "INCORPORATED", "LLC", "LLP", "LP", "LTD", "LIMITED",
    "CORP", "CORPORATION", "CO", "COMPANY", "PC", "PLLC",
}

AUTO_MATCH_THRESHOLD = 0.90
REVIEW_MATCH_THRESHOLD = 0.75

'''
Header normalization
'''
def header_to_snake(h: str) -> str:
    # Strip all non-alphanumeric chars (M/WBE >> MWBE, % >> gone), then snake_case
    clean = re.sub(r'[^a-zA-Z0-9 ]', '', h)
    return '_'.join(p.lower() for p in clean.split())

def is_amount_col(col: str) -> bool:
    return "amount" in col or "paid_to_date" in col or "spend_to_date" in col or "spent_to_date" in col


'''
Loaders, both return (rows: list[dict], cols: list[str])
'''

def load_csv_rows(path: str, label: str) -> tuple:
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        snake = {h: header_to_snake(h) for h in (reader.fieldnames or []) if h is not None}
        for raw in reader:
            rows.append({snake[k]: (v or "").strip() for k, v in raw.items() if k in snake})
    print(f"[{label}] Loaded {len(rows)} rows from {path}")
    return rows, list(snake.values())


def load_excel_rows(path: str, label: str, max_search: int = 20) -> tuple:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Find header row as the one with the most non-None string cells
    best_idx, best_row, best_count = None, None, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True)):
        n_strings = sum(1 for v in row if isinstance(v, str) and v.strip())
        if n_strings > best_count:
            best_count, best_row, best_idx = n_strings, row, i + 1

    if best_count < 2:
        print(f"[{label}] WARNING: no header row found in {path}")
        wb.close()
        return [], []

    # Map column index >> snake name
    col_positions = {j: header_to_snake(str(v)) for j, v in enumerate(best_row) if v is not None}

    rows = []
    for row in ws.iter_rows(min_row=best_idx + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        record = {name: (str(row[j]).strip() if j < len(row) and row[j] is not None else "")
                  for j, name in col_positions.items()}
        rows.append(record)

    wb.close()
    print(f"[{label}] Loaded {len(rows)} rows from {path}")
    return rows, list(col_positions.values())


def load_table(path: str, table_name: str, label: str = None) -> tuple:
    label = label or table_name
    if path.lower().endswith(".xlsx"):
        return load_excel_rows(path, label)
    return load_csv_rows(path, label)


'''
Database
'''


def create_raw_table(conn: sqlite3.Connection, table_name: str, cols: list) -> None:
    extra = TABLE_CONFIGS.get(table_name, [])
    col_defs = ",\n        ".join(f"{c} REAL" if is_amount_col(c) else f"{c} TEXT" for c in cols)
    extra_defs = (",\n        " + ",\n        ".join(extra)) if extra else ""
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(f"CREATE TABLE {table_name} (\n        {col_defs}{extra_defs}\n    )")


def create_derived_tables(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    DROP TABLE IF EXISTS contracts_unified;
    DROP TABLE IF EXISTS vendor_summary;

    CREATE TABLE contracts_unified (
        contract_id     TEXT,
        vendor_name     TEXT,
        vendor_role     TEXT,
        mwbe_category   TEXT,
        purpose         TEXT,
        current_amount  REAL,
        original_amount REAL,
        award_method    TEXT,
        contract_type   TEXT,
        start_date      TEXT,
        end_date        TEXT,
        status          TEXT,
        is_diit         INTEGER DEFAULT 0
    );

    CREATE TABLE vendor_summary (
        vendor_name     TEXT PRIMARY KEY,
        num_contracts   INTEGER,
        total_amount    REAL,
        mwbe_category   TEXT,
        pct_of_total    REAL
    );
    """)
    conn.commit()


def create_ownership_tables(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    DROP TABLE IF EXISTS vendor_passport_link;
    DROP TABLE IF EXISTS ownership_cluster;
    DROP TABLE IF EXISTS cluster_summary;
    DROP TABLE IF EXISTS address_flag;

    CREATE TABLE vendor_passport_link (
        checkbook_vendor_name TEXT PRIMARY KEY,
        passport_vendor_name  TEXT,
        match_score           REAL,
        match_type            TEXT
    );

    CREATE TABLE ownership_cluster (
        vendor_name TEXT PRIMARY KEY,
        cluster_id  TEXT
    );

    CREATE TABLE cluster_summary (
        cluster_id     TEXT PRIMARY KEY,
        member_count   INTEGER,
        member_vendors TEXT,
        num_contracts  INTEGER,
        total_amount   REAL,
        pct_of_total   REAL
    );

    CREATE TABLE address_flag (
        address_key TEXT,
        vendor_name TEXT
    );
    """)
    conn.commit()


def parse_amount(val: str) -> float:
    if not val:
        return 0.0
    val = val.replace(",", "").replace("$", "").strip()
    if val in ("", "-"):
        return 0.0
    try:
        return float(val)
    except ValueError:
        return 0.0


def insert_rows(conn: sqlite3.Connection, table_name: str, rows: list) -> None:
    if not rows:
        return

    cols = list(rows[0].keys())
    placeholders = ", ".join("?" for _ in cols)
    sql = f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({placeholders})"
    values = [
        tuple(parse_amount(r.get(c, "")) if is_amount_col(c) else r.get(c, "") for c in cols)
        for r in rows
    ]
    conn.executemany(sql, values)
    conn.commit()


'''
DIIT flagging, two-pass SQL on raw tables
'''

def _like_clause(column: str, terms: list) -> tuple:
    clause = " OR ".join(f"{column} LIKE ?" for _ in terms)
    params = [f"%{t}%" for t in terms]
    return clause, params


def flag_diit_sql(conn: sqlite3.Connection, loaded: set) -> None:
    if "contracts_registered" in loaded:
        prime_kw, prime_kw_params = _like_clause("prime_contract_purpose", DIIT_KEYWORDS)
        sub_kw, sub_kw_params = _like_clause("sub_contract_purpose", DIIT_KEYWORDS)
        conn.execute(
            f"UPDATE contracts_registered SET is_diit=1 WHERE {prime_kw} OR {sub_kw}",
            prime_kw_params + sub_kw_params,
        )
        prime_excl, prime_excl_params = _like_clause("prime_contract_purpose", DIIT_EXCLUDE_PHRASES)
        sub_excl, sub_excl_params = _like_clause("sub_contract_purpose", DIIT_EXCLUDE_PHRASES)
        conn.execute(
            f"UPDATE contracts_registered SET is_diit=0 WHERE is_diit=1 AND ({prime_excl} OR {sub_excl})",
            prime_excl_params + sub_excl_params,
        )

    if "contracts_pending" in loaded:
        pend_kw, pend_kw_params = _like_clause("purpose", DIIT_KEYWORDS)
        conn.execute(f"UPDATE contracts_pending SET is_diit=1 WHERE {pend_kw}", pend_kw_params)
        pend_excl, pend_excl_params = _like_clause("purpose", DIIT_EXCLUDE_PHRASES)
        conn.execute(
            f"UPDATE contracts_pending SET is_diit=0 WHERE is_diit=1 AND ({pend_excl})",
            pend_excl_params,
        )

    conn.commit()


'''
Analysis, unified table, vendor summary, HHI
'''

def build_unified_table(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM contracts_unified")

    conn.execute("""
        INSERT INTO contracts_unified
            (contract_id, vendor_name, vendor_role, mwbe_category, purpose,
             current_amount, original_amount, award_method, contract_type,
             start_date, end_date, status, is_diit)
        SELECT
            prime_contract_id,
            MAX(prime_vendor),
            'prime',
            MAX(prime_vendor_mwbe_category),
            MAX(prime_contract_purpose),
            MAX(prime_contract_current_amount),
            MAX(prime_contract_original_amount),
            MAX(prime_contract_award_method),
            MAX(prime_contract_type),
            MAX(prime_contract_start_date),
            MAX(prime_contract_end_date),
            'registered',
            MAX(is_diit)
        FROM contracts_registered
        WHERE prime_vendor IS NOT NULL AND prime_vendor != ''
        GROUP BY prime_contract_id
    """)

    #Sub-vendor INSERT removed: '-' placeholder leaked 40k phantom $0 rows
    conn.execute("""
        INSERT INTO contracts_unified
            (contract_id, vendor_name, vendor_role, mwbe_category, purpose,
             current_amount, original_amount, award_method, contract_type,
             start_date, end_date, status, is_diit)
        SELECT
            contract_id, prime_vendor, 'prime', prime_mwbe_category, purpose,
            current_amount, original_amount, award_method, contract_type,
            start_date, end_date, 'pending', is_diit
        FROM contracts_pending
        WHERE prime_vendor IS NOT NULL AND prime_vendor != ''
    """)
    conn.commit()


def build_vendor_summary_sql(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM vendor_summary")
    conn.execute("""
        INSERT INTO vendor_summary (vendor_name, num_contracts, total_amount, mwbe_category, pct_of_total)
        SELECT
            vendor_name,
            COUNT(DISTINCT contract_id) AS num_contracts,
            SUM(current_amount)         AS total_amount,
            (SELECT mwbe_category FROM contracts_unified u2
             WHERE u2.vendor_name = u1.vendor_name AND u2.is_diit = 1
             GROUP BY mwbe_category ORDER BY COUNT(*) DESC LIMIT 1) AS mwbe_category,
            0.0
        FROM contracts_unified u1
        WHERE is_diit = 1
        GROUP BY vendor_name
    """)
    conn.commit()
    total = conn.execute("SELECT SUM(total_amount) FROM vendor_summary").fetchone()[0] or 0
    if total > 0:
        conn.execute("UPDATE vendor_summary SET pct_of_total = ROUND(total_amount * 100.0 / ?, 2)", (total,))
        conn.commit()


def compute_hhi(conn: sqlite3.Connection) -> float:
    rows = conn.execute("SELECT pct_of_total FROM vendor_summary").fetchall()
    return sum(r[0] ** 2 for r in rows if r[0] is not None)


'''
PASSPort name matching and ownership clustering
'''

def normalize_name(name: str) -> str:
    name = name.upper()
    name = name.replace("&", " AND ")
    name = re.sub(r"[.,'/\-]", " ", name)
    tokens = name.split()
    while tokens and tokens[-1] in GENERIC_SUFFIXES:
        tokens.pop()
    return " ".join(tokens)


def normalize_address(line1: str, zip_code: str) -> str:
    line1 = re.sub(r"[^A-Z0-9 ]", "", (line1 or "").upper())
    line1 = " ".join(line1.split())
    zip_code = (zip_code or "").strip()[:5]
    return f"{line1}|{zip_code}"


def build_block_index(names):
    index = {}
    for name in names:
        norm = normalize_name(name)
        tokens = norm.split()
        if not tokens:
            continue
        index.setdefault(tokens[0], []).append((norm, name))
    return index


def best_match(name, block_index):
    norm = normalize_name(name)
    tokens = norm.split()
    if not tokens:
        return None, 0.0
    candidates = block_index.get(tokens[0], [])
    best_name, best_score = None, 0.0
    for candidate_norm, candidate_name in candidates:
        score = SequenceMatcher(None, norm, candidate_norm).ratio()
        if score > best_score:
            best_score = score
            best_name = candidate_name
    return best_name, best_score


def classify_score(score):
    if score >= 0.999:
        return "exact"
    if score >= AUTO_MATCH_THRESHOLD:
        return "fuzzy_auto"
    if score >= REVIEW_MATCH_THRESHOLD:
        return "fuzzy_review"
    return "unmatched"


def link_checkbook_to_passport(conn):
    checkbook_vendors = [r[0] for r in conn.execute("SELECT vendor_name FROM vendor_summary")]
    passport_vendors = [r[0] for r in conn.execute(
        "SELECT DISTINCT vendor_name FROM completeentityprincipalwebsites"
    )]
    passport_index = build_block_index(passport_vendors)

    links = []
    for cb_name in checkbook_vendors:
        match_name, score = best_match(cb_name, passport_index)
        match_type = classify_score(score) if match_name else "unmatched"
        links.append((cb_name, match_name if match_type != "unmatched" else None, score, match_type))

    conn.executemany(
        "INSERT INTO vendor_passport_link VALUES (?, ?, ?, ?)",
        links
    )
    conn.commit()
    return links


class UnionFind:
    def __init__(self, items):
        self.parent = {item: item for item in items}

    def find(self, item):
        while self.parent[item] != item:
            self.parent[item] = self.parent[self.parent[item]]
            item = self.parent[item]
        return item

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[ra] = rb


def cluster_by_shared_principals(conn, uf, confirmed_links):
    passport_to_principals = {}
    for passport_name, principal_name in conn.execute(
        "SELECT vendor_name, principal_name FROM completeentityprincipalwebsites"
    ):
        if principal_name and principal_name.strip():
            passport_to_principals.setdefault(passport_name, set()).add(principal_name.strip().upper())

    vendor_to_principals = {
        cb_name: passport_to_principals.get(passport_name, set())
        for cb_name, passport_name in confirmed_links.items()
    }

    principal_to_vendors = {}
    for cb_name, principals in vendor_to_principals.items():
        for principal in principals:
            principal_to_vendors.setdefault(principal, set()).add(cb_name)

    shared_principal_pairs = []
    for principal, vendors in principal_to_vendors.items():
        if len(vendors) > 1:
            vendor_list = sorted(vendors)
            first = vendor_list[0]
            for other in vendor_list[1:]:
                uf.union(first, other)
                shared_principal_pairs.append((principal, first, other))

    return shared_principal_pairs


def cluster_by_related_entities(conn, uf, confirmed_links):
    checkbook_index = build_block_index(list(confirmed_links.keys()))

    passport_to_related = {}
    for passport_name, related_name in conn.execute(
        "SELECT vendor_name, related_entity_name FROM completeentityrelatedentitieswebsite"
    ):
        if related_name and related_name.strip():
            passport_to_related.setdefault(passport_name, set()).add(related_name)

    related_pairs = []
    for cb_name, passport_name in confirmed_links.items():
        for related_name in passport_to_related.get(passport_name, ()):
            match_name, score = best_match(related_name, checkbook_index)
            if match_name and classify_score(score) in ("exact", "fuzzy_auto") and match_name != cb_name:
                uf.union(cb_name, match_name)
                related_pairs.append((cb_name, related_name, match_name, score))

    return related_pairs


def flag_shared_addresses(conn, confirmed_links):
    passport_to_address = {}
    for passport_name, line1, zip_code in conn.execute(
        "SELECT vendor_name, address_line_1, zip_code FROM completeentitysummarywebsite"
    ):
        passport_to_address.setdefault(passport_name, (line1, zip_code))

    address_to_vendors = {}
    for cb_name, passport_name in confirmed_links.items():
        row = passport_to_address.get(passport_name)
        if not row:
            continue
        key = normalize_address(row[0], row[1])
        if key == "|":
            continue
        address_to_vendors.setdefault(key, set()).add(cb_name)

    flagged = {k: v for k, v in address_to_vendors.items() if len(v) > 1}
    rows = [(key, vendor) for key, vendors in flagged.items() for vendor in vendors]
    if rows:
        conn.executemany("INSERT INTO address_flag VALUES (?, ?)", rows)
        conn.commit()
    return flagged


def build_cluster_summary(conn, uf):
    vendor_amounts = {
        r[0]: (r[1], r[2]) for r in conn.execute(
            "SELECT vendor_name, num_contracts, total_amount FROM vendor_summary"
        )
    }

    cluster_members = {}
    for vendor_name in vendor_amounts:
        root = uf.find(vendor_name)
        cluster_members.setdefault(root, []).append(vendor_name)

    conn.execute("DELETE FROM ownership_cluster")
    conn.execute("DELETE FROM cluster_summary")

    grand_total = sum(amt for _, amt in vendor_amounts.values())

    cluster_rows = []
    ownership_rows = []
    for cluster_id, members in cluster_members.items():
        num_contracts = sum(vendor_amounts[m][0] for m in members)
        total_amount = sum(vendor_amounts[m][1] for m in members)
        pct_of_total = round(total_amount * 100.0 / grand_total, 2) if grand_total else 0.0
        cluster_rows.append((
            cluster_id, len(members), "; ".join(sorted(members)),
            num_contracts, total_amount, pct_of_total
        ))
        for m in members:
            ownership_rows.append((m, cluster_id))

    conn.executemany("INSERT INTO cluster_summary VALUES (?, ?, ?, ?, ?, ?)", cluster_rows)
    conn.executemany("INSERT INTO ownership_cluster VALUES (?, ?)", ownership_rows)
    conn.commit()


def compute_cluster_hhi(conn):
    rows = conn.execute("SELECT pct_of_total FROM cluster_summary").fetchall()
    return sum(r[0] ** 2 for r in rows if r[0] is not None)


def build_market_share_chart(conn, output_path, top_n=10):
    rows = conn.execute(
        "SELECT cluster_id, total_amount, pct_of_total FROM cluster_summary ORDER BY total_amount DESC"
    ).fetchall()
    total = sum(r[1] for r in rows)
    top_rows = rows[:top_n]
    other_pct = 100 * (total - sum(r[1] for r in top_rows)) / total if total else 0
    other_n = len(rows) - len(top_rows)

    labels = [r[0][:35] for r in top_rows] + [f"All others ({other_n} vendors)"]
    pcts = [r[2] for r in top_rows] + [other_pct]
    colors = ["#2a78d6"] * len(top_rows) + ["#c3c2b7"]

    fig, ax = plt.subplots(figsize=(8, 6))
    y_pos = range(len(labels))
    ax.barh(y_pos, pcts, color=colors)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_xlabel("% of total DIIT contract dollars")
    ax.set_title("Share of DIIT contract dollars by vendor/ownership cluster")
    ax.grid(axis="x", alpha=0.3)
    common.save_chart(fig, output_path)


def run_ownership_clustering(conn, figures_dir="."):
    create_ownership_tables(conn)

    links = link_checkbook_to_passport(conn)
    match_counts = {}
    for _, _, _, match_type in links:
        match_counts[match_type] = match_counts.get(match_type, 0) + 1

    print("\n--- Vendor -> PASSPort profile linking ---")
    for match_type in ("exact", "fuzzy_auto", "fuzzy_review", "unmatched"):
        print(f"  {match_type:<14} {match_counts.get(match_type, 0)}")

    confirmed_links = {
        cb_name: passport_name for cb_name, passport_name, score, match_type in links
        if match_type in ("exact", "fuzzy_auto")
    }
    review_links = [
        (cb_name, passport_name, score) for cb_name, passport_name, score, match_type in links
        if match_type == "fuzzy_review"
    ]

    uf = UnionFind([r[0] for r in conn.execute("SELECT vendor_name FROM vendor_summary")])

    shared_principal_pairs = cluster_by_shared_principals(conn, uf, confirmed_links)
    related_pairs = cluster_by_related_entities(conn, uf, confirmed_links)
    address_flags = flag_shared_addresses(conn, confirmed_links)

    build_cluster_summary(conn, uf)
    cluster_hhi = compute_cluster_hhi(conn)
    baseline_hhi = compute_hhi(conn)

    print(f"\nBaseline name-string HHI: {baseline_hhi:.1f}")
    print(f"Ownership-adjusted HHI:  {cluster_hhi:.1f}")

    multi_member_clusters = conn.execute(
        "SELECT cluster_id, member_count, member_vendors, total_amount, pct_of_total "
        "FROM cluster_summary WHERE member_count > 1 ORDER BY total_amount DESC"
    ).fetchall()

    print(f"\n--- Ownership clusters with 2+ vendors ({len(multi_member_clusters)}) ---")
    for cluster_id, count, members, total_amount, pct in multi_member_clusters:
        print(f"  [{count}] {members}  ${total_amount:,.2f}  {pct}%")

    if shared_principal_pairs:
        print(f"\n--- Merges from shared principal name ({len(shared_principal_pairs)}) ---")
        for principal, a, b in shared_principal_pairs:
            print(f"  {a}  <->  {b}   (shared principal: {principal})")

    if related_pairs:
        print(f"\n--- Merges from self-reported related entity ({len(related_pairs)}) ---")
        for cb_name, related_name, matched_vendor, score in related_pairs:
            print(f"  {cb_name}  ->  related entity \"{related_name}\"  matched to  {matched_vendor}  (score={score:.2f})")

    if review_links:
        print(f"\n--- Needs manual review ({len(review_links)}) ---")
        for cb_name, passport_name, score in sorted(review_links, key=lambda r: -r[2]):
            print(f"  {cb_name}  ~  {passport_name}   (score={score:.2f})")

    if address_flags:
        print(f"\n--- Vendors sharing a registered address ({len(address_flags)} address groups, informational only) ---")
        for key, vendors in address_flags.items():
            line1 = key.split("|")[0]
            print(f"  {line1}: {', '.join(sorted(vendors))}")

    chart_path = f"{figures_dir}/market_share_by_vendor.png"
    build_market_share_chart(conn, chart_path)
    print(f"\nSaved chart: {chart_path}")


'''
Main
'''

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Checkbook NYC + PASSPort exports >> SQLite >> DIIT contract analysis + ownership clustering."
    )
    parser.add_argument("--registered",      help="Checkbook registered contracts CSV")
    parser.add_argument("--pending",         help="Checkbook pending contracts CSV")
    parser.add_argument("--sources", nargs="+", metavar="FILE",
                        help="Any additional source files (CSV or XLSX). Table name derived from filename.")
    common.add_db_figures_args(parser, db_help="Output SQLite database path")
    args = parser.parse_args()
    common.ensure_figures_dir(args.figures_dir)

    if not args.registered and not args.pending:
        parser.error("Provide at least --registered or --pending.")

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA foreign_keys = ON")
    create_derived_tables(conn)

    # Load each file, build its table dynamically from actual columns, then insert
    sources = []
    if args.registered:
        sources.append(("contracts_registered", args.registered, "registered"))
    if args.pending:
        sources.append(("contracts_pending", args.pending, "pending"))
    for path in (args.sources or []):
        # Table name comes from filename
        stem = os.path.splitext(os.path.basename(path))[0]
        table_name = header_to_snake(stem)
        sources.append((table_name, path, table_name))

    loaded = set()
    for table_name, path, label in sources:
        rows, cols = load_table(path, table_name, label)
        create_raw_table(conn, table_name, cols)
        insert_rows(conn, table_name, rows)
        loaded.add(table_name)

    reg_n  = conn.execute("SELECT COUNT(*) FROM contracts_registered").fetchone()[0] if "contracts_registered" in loaded else 0
    pend_n = conn.execute("SELECT COUNT(*) FROM contracts_pending").fetchone()[0]    if "contracts_pending"    in loaded else 0
    print(f"\nLoaded {reg_n:,} registered and {pend_n} pending rows into {args.db}")

    flag_diit_sql(conn, loaded)
    build_unified_table(conn)

    diit_count = conn.execute("SELECT COUNT(*) FROM contracts_unified WHERE is_diit=1").fetchone()[0]
    print(f"Flagged {diit_count} unified rows as likely DIIT/tech contracts.")

    build_vendor_summary_sql(conn)
    hhi = compute_hhi(conn)

    print(f"\nHHI (DIIT-flagged vendors): {hhi:.1f}")
    bands = [(1500, "Unconcentrated"), (2500, "Moderately concentrated"), (float("inf"), "Highly concentrated")]
    print(f"Interpretation (DOJ/FTC): {next(label for cap, label in bands if hhi < cap)}")

    print("\n--- Top 10 vendors by DIIT contract value ---")
    for row in conn.execute("""
        SELECT vendor_name, num_contracts, total_amount, mwbe_category, pct_of_total
        FROM vendor_summary ORDER BY total_amount DESC LIMIT 10
    """).fetchall():
        print(f"  {row[0]:<40} contracts={row[1]:<4} ${row[2]:,.2f}  {row[3]:<15} {row[4]}%")

    print("\n--- M/WBE breakdown (DIIT-flagged) ---")
    for cat, cnt in conn.execute("""
        SELECT mwbe_category, COUNT(*) FROM contracts_unified
        WHERE is_diit=1 GROUP BY mwbe_category ORDER BY COUNT(*) DESC
    """).fetchall():
        print(f"  {cat or '(blank)':<22} {cnt}")

    extra_loaded = [t for t in loaded if t not in ("contracts_registered", "contracts_pending")]
    if extra_loaded:
        print("\n--- Additional source tables loaded ---")
        for t in extra_loaded:
            n = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            print(f"  {t}: {n:,} rows")

    has_passport_data = any(t == "completeentityprincipalwebsites" for t in loaded)
    if has_passport_data:
        run_ownership_clustering(conn, figures_dir=args.figures_dir)
    else:
        print("\nNo completeentityprincipalwebsites source loaded - skipping ownership clustering.")

    conn.close()
    print(f"\nDone. Database: {args.db}")
    print("Query with: sqlite3 " + args.db)