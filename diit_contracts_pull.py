"""
Bryant Luna-Ramos
6/16/26

How to run:
    python diit_contracts_pull.py --registered registered.csv --pending pending.csv
    python diit_contracts_pull.py --registered registered.csv --pending pending.csv \
        --sources complete_entity_principal_websites.xlsx \
                  complete_entity_relatedentities_website.xlsx \
                  complete_entity_othernames_website.xlsx \
                  complete_entity_summary_website.xlsx \
"""

import argparse
import csv
import openpyxl
import re
import sqlite3

DB_PATH = "diit_contracts.db"

DIIT_KEYWORDS = [
    "DIIT", "information technology", "instructional technology", "technology",
    "software", "hardware", "network", "server", "laptop", "desktop", "tablet",
    "chromebook", "wireless", "data center", "cabling", "IT services",
    "IT support", "IT consulting", "cloud", "digital", "computer", "device", "cyber",
    "telecommunications", "infrastructure", "system",
]

DIIT_EXCLUDE_PHRASES = [
    "family child care",
    "crisis management system",
    "system-wide", "systemwide", "system wide",
    "fire alarm", "fire suppression", "sprinkler", "standpipe",
    "security system",
    "hvac", "air condition", "boiler", "plumbing", "backflow", "fuel oil",
    "public address system", "gas leak detection", "de-watering",
    "kitchen exhaust", "water treatment", "direct digital control",
    "window shades",
    "legal process server",
    "vendor does not have order in system", "doc posted in city",
]

# Extra columns appended to Checkbook tables beyond what the source file provides
# All other tables (PASSPort and any future source) get no extra columns by default
TABLE_CONFIGS = {
    "contracts_registered": ["is_diit INTEGER DEFAULT 0"],
    "contracts_pending":    ["is_diit INTEGER DEFAULT 0"],
}

'''
Header normalization
'''
def header_to_snake(h: str) -> str:
    # Strip all non-alphanumeric chars (M/WBE >> MWBE, % >> gone), then snake_case
    clean = re.sub(r'[^a-zA-Z0-9 ]', '', h)
    return '_'.join(p.lower() for p in clean.split())

def is_amount_col(col: str) -> bool:
    return "amount" in col or "paid_to_date" in col or "spend_to_date" in col or "spent_to_date" in col



''' 
Loaders, both return (rows: list[dict], cols: list[str])
'''


def load_csv_rows(path: str, label: str) -> tuple:
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        snake = {h: header_to_snake(h) for h in (reader.fieldnames or []) if h is not None}
        for raw in reader:
            rows.append({snake[k]: (v or "").strip() for k, v in raw.items() if k in snake})
    print(f"[{label}] Loaded {len(rows)} rows from {path}")
    return rows, list(snake.values())


def load_excel_rows(path: str, label: str, max_search: int = 20) -> tuple:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Find header row as the one with the most non-None string cells
    best_idx, best_row, best_count = None, None, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True)):
        n_strings = sum(1 for v in row if isinstance(v, str) and v.strip())
        if n_strings > best_count:
            best_count, best_row, best_idx = n_strings, row, i + 1

    if best_count < 2:
        print(f"[{label}] WARNING: no header row found in {path}")
        wb.close()
        return [], []

    # Map column index >> snake name
    col_positions = {j: header_to_snake(str(v)) for j, v in enumerate(best_row) if v is not None}

    rows = []
    for row in ws.iter_rows(min_row=best_idx + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        record = {name: (str(row[j]).strip() if j < len(row) and row[j] is not None else "")
                  for j, name in col_positions.items()}
        rows.append(record)

    wb.close()
    print(f"[{label}] Loaded {len(rows)} rows from {path}")
    return rows, list(col_positions.values())


def load_table(path: str, table_name: str, label: str = None) -> tuple:
    # Single entry point, dispatches to CSV or Excel loader by extension
    label = label or table_name
    if path.lower().endswith(".xlsx"):
        return load_excel_rows(path, label)
    return load_csv_rows(path, label)


'''
Database
'''

def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def create_raw_table(conn: sqlite3.Connection, table_name: str, cols: list) -> None:
    extra = TABLE_CONFIGS.get(table_name, [])
    col_defs = ",\n        ".join(f"{c} REAL" if is_amount_col(c) else f"{c} TEXT" for c in cols)
    extra_defs = (",\n        " + ",\n        ".join(extra)) if extra else ""
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(f"CREATE TABLE {table_name} (\n        {col_defs}{extra_defs}\n    )")


def create_derived_tables(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    DROP TABLE IF EXISTS contracts_unified;
    DROP TABLE IF EXISTS vendor_summary;

    CREATE TABLE contracts_unified (
        contract_id     TEXT,
        vendor_name     TEXT,
        vendor_role     TEXT,
        mwbe_category   TEXT,
        purpose         TEXT,
        current_amount  REAL,
        original_amount REAL,
        award_method    TEXT,
        contract_type   TEXT,
        start_date      TEXT,
        end_date        TEXT,
        status          TEXT,
        is_diit         INTEGER DEFAULT 0
    );

    CREATE TABLE vendor_summary (
        vendor_name     TEXT PRIMARY KEY,
        num_contracts   INTEGER,
        total_amount    REAL,
        mwbe_category   TEXT,
        pct_of_total    REAL
    );
    """)
    conn.commit()


def parse_amount(val: str) -> float:
    if not val:
        return 0.0
    val = val.replace(",", "").replace("$", "").strip()
    if val in ("", "-"):
        return 0.0
    try:
        return float(val)
    except ValueError:
        return 0.0


def insert_rows(conn: sqlite3.Connection, table_name: str, rows: list) -> None:
    if not rows:
        return
    # Column order comes from the row dicts themselves, no separate column list needed
    cols = list(rows[0].keys())
    placeholders = ", ".join("?" for _ in cols)
    sql = f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({placeholders})"
    values = [
        tuple(parse_amount(r.get(c, "")) if is_amount_col(c) else r.get(c, "") for c in cols)
        for r in rows
    ]
    conn.executemany(sql, values)
    conn.commit()


'''
DIIT flagging, two-pass SQL on raw tables
'''


def flag_diit_sql(conn: sqlite3.Connection, loaded: set) -> None:
    if "contracts_registered" in loaded:
        like_prime = " OR ".join(f"prime_contract_purpose LIKE '%{kw}%'" for kw in DIIT_KEYWORDS)
        like_sub   = " OR ".join(f"sub_contract_purpose LIKE '%{kw}%'" for kw in DIIT_KEYWORDS)
        conn.execute(f"UPDATE contracts_registered SET is_diit=1 WHERE {like_prime} OR {like_sub}")
        excl_prime = " OR ".join(f"prime_contract_purpose LIKE '%{p}%'" for p in DIIT_EXCLUDE_PHRASES)
        excl_sub   = " OR ".join(f"sub_contract_purpose LIKE '%{p}%'" for p in DIIT_EXCLUDE_PHRASES)
        conn.execute(f"UPDATE contracts_registered SET is_diit=0 WHERE is_diit=1 AND ({excl_prime} OR {excl_sub})")

    if "contracts_pending" in loaded:
        like_pend = " OR ".join(f"purpose LIKE '%{kw}%'" for kw in DIIT_KEYWORDS)
        conn.execute(f"UPDATE contracts_pending SET is_diit=1 WHERE {like_pend}")
        excl_pend = " OR ".join(f"purpose LIKE '%{p}%'" for p in DIIT_EXCLUDE_PHRASES)
        conn.execute(f"UPDATE contracts_pending SET is_diit=0 WHERE is_diit=1 AND ({excl_pend})")

    conn.commit()


'''
 Analysis, unified table, vendor summary, HHI
'''

def build_unified_table(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM contracts_unified")

    conn.execute("""
        INSERT INTO contracts_unified
            (contract_id, vendor_name, vendor_role, mwbe_category, purpose,
             current_amount, original_amount, award_method, contract_type,
             start_date, end_date, status, is_diit)
        SELECT
            prime_contract_id, prime_vendor, 'prime', prime_vendor_mwbe_category,
            prime_contract_purpose, prime_contract_current_amount,
            prime_contract_original_amount, prime_contract_award_method,
            prime_contract_type, prime_contract_start_date, prime_contract_end_date,
            'registered', is_diit
        FROM contracts_registered
        WHERE prime_vendor IS NOT NULL AND prime_vendor != ''
    """)

    #Sub-vendor INSERT removed: '-' placeholder leaked 40k phantom $0 rows

    conn.execute("""
        INSERT INTO contracts_unified
            (contract_id, vendor_name, vendor_role, mwbe_category, purpose,
             current_amount, original_amount, award_method, contract_type,
             start_date, end_date, status, is_diit)
        SELECT
            contract_id, prime_vendor, 'prime', prime_mwbe_category, purpose,
            current_amount, original_amount, award_method, contract_type,
            start_date, end_date, 'pending', is_diit
        FROM contracts_pending
        WHERE prime_vendor IS NOT NULL AND prime_vendor != ''
    """)
    conn.commit()


def build_vendor_summary_sql(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM vendor_summary")
    conn.execute("""
        INSERT INTO vendor_summary (vendor_name, num_contracts, total_amount, mwbe_category, pct_of_total)
        SELECT
            vendor_name,
            COUNT(DISTINCT contract_id) AS num_contracts,
            SUM(current_amount)         AS total_amount,
            (SELECT mwbe_category FROM contracts_unified u2
             WHERE u2.vendor_name = u1.vendor_name AND u2.is_diit = 1
             GROUP BY mwbe_category ORDER BY COUNT(*) DESC LIMIT 1) AS mwbe_category,
            0.0
        FROM contracts_unified u1
        WHERE is_diit = 1
        GROUP BY vendor_name
    """)
    conn.commit()
    total = conn.execute("SELECT SUM(total_amount) FROM vendor_summary").fetchone()[0] or 0
    if total > 0:
        conn.execute("UPDATE vendor_summary SET pct_of_total = ROUND(total_amount * 100.0 / ?, 2)", (total,))
        conn.commit()


def compute_hhi(conn: sqlite3.Connection) -> float:
    # HHI = sum of squared market shares (0-10000 scale)
    rows = conn.execute("SELECT pct_of_total FROM vendor_summary").fetchall()
    return sum(r[0] ** 2 for r in rows if r[0] is not None)



''' 
Mains
'''


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Checkbook NYC + PASSPort exports >> SQLite >> DIIT contract analysis."
    )
    parser.add_argument("--registered",      help="Checkbook registered contracts CSV")
    parser.add_argument("--pending",         help="Checkbook pending contracts CSV")
    parser.add_argument("--sources", nargs="+", metavar="FILE",
                        help="Any additional source files (CSV or XLSX). Table name derived from filename.")
    args = parser.parse_args()

    if not args.registered and not args.pending:
        parser.error("Provide at least --registered or --pending.")

    conn = get_connection()
    create_derived_tables(conn)

    # Load each file, build its table dynamically from actual columns, then insert
    # Build source list: fixed Checkbook files + arbitrary extra sources
    import os
    sources = []
    if args.registered:
        sources.append(("contracts_registered", args.registered, "registered"))
    if args.pending:
        sources.append(("contracts_pending", args.pending, "pending"))
    for path in (args.sources or []):
        # Table name comes from filename
        stem = os.path.splitext(os.path.basename(path))[0]
        table_name = header_to_snake(stem)
        sources.append((table_name, path, table_name))

    loaded = set()
    for table_name, path, label in sources:
        rows, cols = load_table(path, table_name, label)
        create_raw_table(conn, table_name, cols)
        insert_rows(conn, table_name, rows)
        loaded.add(table_name)

    reg_n  = conn.execute("SELECT COUNT(*) FROM contracts_registered").fetchone()[0] if "contracts_registered" in loaded else 0
    pend_n = conn.execute("SELECT COUNT(*) FROM contracts_pending").fetchone()[0]    if "contracts_pending"    in loaded else 0
    print(f"\nLoaded {reg_n:,} registered and {pend_n} pending rows into {DB_PATH}")

    flag_diit_sql(conn, loaded)
    build_unified_table(conn)

    diit_count = conn.execute("SELECT COUNT(*) FROM contracts_unified WHERE is_diit=1").fetchone()[0]
    print(f"Flagged {diit_count} unified rows as likely DIIT/tech contracts.")

    build_vendor_summary_sql(conn)
    hhi = compute_hhi(conn)

    print(f"\nHHI (DIIT-flagged vendors): {hhi:.1f}")
    bands = [(1500, "Unconcentrated"), (2500, "Moderately concentrated"), (float("inf"), "Highly concentrated")]
    print(f"Interpretation (DOJ/FTC): {next(label for cap, label in bands if hhi < cap)}")

    print("\n--- Top 10 vendors by DIIT contract value ---")
    for row in conn.execute("""
        SELECT vendor_name, num_contracts, total_amount, mwbe_category, pct_of_total
        FROM vendor_summary ORDER BY total_amount DESC LIMIT 10
    """).fetchall():
        print(f"  {row[0]:<40} contracts={row[1]:<4} ${row[2]:,.2f}  {row[3]:<15} {row[4]}%")

    print("\n--- M/WBE breakdown (DIIT-flagged) ---")
    for cat, cnt in conn.execute("""
        SELECT mwbe_category, COUNT(*) FROM contracts_unified
        WHERE is_diit=1 GROUP BY mwbe_category ORDER BY COUNT(*) DESC
    """).fetchall():
        print(f"  {cat or '(blank)':<22} {cnt}")

    extra_loaded = [t for t in loaded if t not in ("contracts_registered", "contracts_pending")]
    if extra_loaded:
        print("\n--- Additional source tables loaded ---")
        for t in extra_loaded:
            n = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            print(f"  {t}: {n:,} rows")

    conn.close()
    print(f"\nDone. Database: {DB_PATH}")
    print("Query with: sqlite3 diit_contracts.db")